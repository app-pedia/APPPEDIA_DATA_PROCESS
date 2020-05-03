import openpyxl
import requests
import json

wb = openpyxl.load_workbook("/Users/ilya/Desktop/google(ar).xlsx.xlsx")
sheet = wb['Sheet1']
column=1
row=1
count=0
#print(sheet.cell(row=1, column=1).value)
#print("%s" % sheet.max_column)
col = sheet['A']               # 첫 번째 열 전체를 가져옴
for cell in col:# 각 행에 대해서

        s = requests.session()
        base_url ='https://www.androidrank.org/api'+str(sheet.cell(row,column).value)
        con = s.get(base_url)
        json_data = json.loads(con.text)


        WhoIsData = []
        WhoIsData.append(json_data['category'])
        WhoIsData.append(json_data['developer']['name'])
        WhoIsData.append(json_data['name'])
        WhoIsData.append(json_data['id'])
        WhoIsData.append(json_data['images']['logo'])
        WhoIsData.append(json_data['installs']['current']['count']) #총 다운로드
        WhoIsData.append(json_data['price']['current'])
        WhoIsData.append(json_data['ratings']['current']['count']) #총 레이팅


        print("카테고리:"),print((json_data['category']))
        print("개발자 이름:"),print((json_data['developer']['name']))
        print("앱 이름:"),print((json_data['name']))
        print("아이디:"),print((json_data['id']))
        print("로고:"),print((json_data['images']['logo']))
        print("총 다운로드:"),print((json_data['installs']['current']['count']))
        print("가격:"),print((json_data['price']['current']))
        print("총 레이팅:"),print((json_data['ratings']['current']['count']))
        print('\n')
        print(count)
        count += 1
        row +=2



