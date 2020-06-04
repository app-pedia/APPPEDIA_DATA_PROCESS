import openpyxl
import requests
import json
from urllib.request import urlopen
from bs4 import BeautifulSoup

wb = openpyxl.load_workbook("/Users/ilya/Desktop/all_id(PS).xlsx")
sheet = wb['Sheet1']
column=1
row=1

#print(sheet.cell(row=1, column=1).value)
#print("%s" % sheet.max_column)
col = sheet['A']               # 첫 번째 열 전체를 가져옴
for cell in col:# 각 행에 대해서
        html = urlopen('https://play.google.com'+str(sheet.cell(row,column).value))
        bsObject = BeautifulSoup(html, "html.parser",from_encoding='utf-8')
        paragraph_data = bsObject.find('div', class_='IxB2fe')
        yes = paragraph_data.find_all('div', class_='hAyfc')
        for span in yes:
            print(span.text)
        print('\n')
        row +=1



