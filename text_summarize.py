# -*- encoding: utf-8 -*-
import openpyxl
import requests
import json
from urllib.request import urlopen
from bs4 import BeautifulSoup
import time
import sys
import logging
#logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)

from gensim.summarization import summarize

wb = openpyxl.load_workbook("/Users/ilya/Desktop/all_id(PS).xlsx")
sheet = wb['Sheet1']
column=1
row=1

excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.column_dimensions['B'].width = 100

num = 0

#print(sheet.cell(row=1, column=1).value)
#print("%s" % sheet.max_column)
col = sheet['A']               # 첫 번째 열 전체를 가져옴

for cell in col:

    try:
            html = urlopen('https://play.google.com'+str(sheet.cell(row,column).value))
            bsObject = BeautifulSoup(html, "html.parser", from_encoding='utf-8')
            paragraph_data = bsObject.find('div', jsname='sngebd')
            i=paragraph_data
            new=(i.text)
            end=(summarize(new, ratio=0.4, word_count=None, split=False))
            print(end)
            print("\n")
            num += 1
            excel_sheet.append([num, end])
            cell_A1 = excel_sheet['A1']
            excel_file.save('data.xlsx')
            excel_file.close()
            row += 1
    except:
        row+=1
        continue




