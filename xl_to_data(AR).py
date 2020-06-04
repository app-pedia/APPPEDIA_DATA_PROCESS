import openpyxl
import requests
import json
from urllib.request import urlopen
from boto.dynamodb2 import results
from bs4 import BeautifulSoup
import time
import sys
import csv
from pandas import ExcelWriter
from konlpy.tag import Okt
from collections import Counter




wb = openpyxl.load_workbook("/Users/ilya/Desktop/developer_data.xlsx")
sheet = wb['Sheet1']
column=1
row=1
count=0
#print(sheet.cell(row=1, column=1).value)
#print("%s" % sheet.max_column)
col = sheet['A']               # 첫 번째 열 전체를 가져옴

excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.column_dimensions['B'].width = 100

num = 0

for cell in col:# 각 행에 대해서
        html = urlopen('https://www.androidrank.org/'+str(sheet.cell(row,column).value))
        bsObject = BeautifulSoup(html, "html.parser")
        paragraph_data = bsObject.find_all('td', style='text-align:left;')
        for a in paragraph_data:
            new=(a.find("a")["href"])
            #print('\n')
            print(new)
            num += 1
            excel_sheet.append([num, new])
            cell_A1 = excel_sheet['A1']
            excel_file.save('data.xlsx')
            excel_file.close()
        time.sleep(1.5)
        row +=1
