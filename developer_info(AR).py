import openpyxl
import requests
import json
from urllib.request import urlopen
from bs4 import BeautifulSoup

wb = openpyxl.load_workbook("/Users/ilya/Desktop/link.xlsx")
sheet = wb['Sheet1']
column=1
row=1
count=0
#print(sheet.cell(row=1, column=1).value)
#print("%s" % sheet.max_column)
col = sheet['A']               # 첫 번째 열 전체를 가져옴

for cell in col:# 각 행에 대해서
        html = urlopen('https://www.androidrank.org/'+str(sheet.cell(row,column).value))
        bsObject = BeautifulSoup(html, "html.parser")
        paragraph_data = bsObject.find_all('table', class_='appstat')
        for tr in paragraph_data:
            new=(tr.text)
            print(new)
            print('\n')
            print(count)
            count += 1
        row +=3
