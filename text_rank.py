# -*- encoding: utf-8 -*-
import openpyxl
import requests
import json
from urllib.request import urlopen
from bs4 import BeautifulSoup
import time
import sys
import logging
from konlpy.tag import Okt
from collections import Counter


wb = openpyxl.load_workbook("/Users/ilya/Desktop/all_id(PS).xlsx")
sheet = wb['Sheet1']
column = 1
row = 1

excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.column_dimensions['A'].width = 100
excel_sheet.column_dimensions['B'].width = 100
excel_sheet.column_dimensions['C'].width = 20

num = 0

# print(sheet.cell(row=1, column=1).value)
# print("%s" % sheet.max_column)
col = sheet['A']  # 첫 번째 열 전체를 가져옴
count=0
for cell in col:  # 각 행에 대해서
    url=str(sheet.cell(row, column).value)
    try:
        html = urlopen('https://play.google.com' + url)
        bsObject = BeautifulSoup(html, "html.parser", from_encoding='utf-8')
        paragraph_data = bsObject.find('div', jsname='sngebd')
        data = bsObject.find('a', itemprop='genre')
        i = paragraph_data
        new = i.text
        okt=Okt()
        main_text = data
        main = main_text.text
        print(main)
        file_format = ['예술', '디자인', '자동차', '뷰티', '패션', '옷', '책', '비즈니스', '만화', '비즈니스', '커뮤니케이션', '교육', '엔터테인먼트',
                       '앤터테인먼트', '경제', '금융', '경영', '음식', '헬스', '건강', '운동', '스포츠', '피트니스', '도서', '독서',
                       '인테리어', '라이프스타일', '라이프', '맵', '지도', 'GPS', '네비게이션', '음악', '노래', '의료', '뉴스', '매거진', '잡지',
                       '포토', '사진', '동영상', '영상', '채널', '위치', '생산', '쇼핑', '소셜', '미디어', '여행', '비디오', '게임',
                       '날씨','미용','음성인식','타투','코딩','메모','메모장','필터','카메라','계산기','시계','시간표','배달','음식','영화','카페','중고','쇼핑몰','코로나','딜리버리','알바','교통','대중교통','문자','메세지','캘린더','켈린','메일','티비','TV']
        a = [main]
        for format in file_format:
            if format in new:
                yes= okt.nouns(format)
                a.append(yes)

        excel_sheet.append(['https://play.google.com' + url, main, str(a), new])
        cell_A1 = excel_sheet['A1']
        excel_file.save('data.xlsx')
        excel_file.close()
        print(a)
        print(count)
        count+=1
        row += 1
    except:
        row+=1
        continue




