import openpyxl
import requests
import json
from urllib.request import urlopen
from bs4 import BeautifulSoup
import time


#wb = openpyxl.load_workbook("/Users/ilya/Desktop/link(developer).xlsx")
#sheet = wb['Sheet1']
#column=1
row=1
count=0
#print(sheet.cell(row=1, column=1).value)
#print("%s" % sheet.max_column)
#col = sheet['A']

# 각 행에 대해서
pageNum = 1
count = 1
i=1321
lastPage = int(i)
num=0

excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.column_dimensions['B'].width = 100

while pageNum < lastPage + 1:
    html1 = urlopen('https://www.androidrank.org/developers/ranking?&start='+str(pageNum))
    bsObject = BeautifulSoup(html1, "html.parser")
    paragraph_data = bsObject.find_all('td', style='text-align:left;')
    #html2=urlopen('https://www.androidrank.org/'+str(sheet.cell(row,column).value))
    for a in paragraph_data:
        print(a.find("a")["href"])
        print('\n')
        new = (a.find("a")["href"])
        # print('\n')
        print(new)
        num += 1
        excel_sheet.append([num, new])
        cell_A1 = excel_sheet['A1']
        cell_A1.alignment = openpyxl.styles.Alignment(horizontal="center")
        cell_B1 = excel_sheet['B1']
        cell_B1.alignment = openpyxl.styles.Alignment(horizontal="center")
        excel_file.save('data.xlsx')
        excel_file.close()
    time.sleep(1.2)
    pageNum += 20
    count += 1

